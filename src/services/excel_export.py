"""Exportació de l'schedule a un fitxer Excel (.xlsx).

Una única fulla «Calendari» amb la vista en quadrícula setmana × dia
(mimica del PDF setmanal). Cada dia té sub-columnes per màquina i
files [Header / Abs / Àrea / Màquina / Matí / Tarda / Nit / Revisió].
Cel·les verdes = presencial, blaves = no presencial, grogues =
peonada, gris = festiu. Prefix «T-» als slots flipats (NP→PRES
forçats)."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from src.domain.constants import GUARDS_RESERVED_SLOT_IDS
from src.domain.schedule_format import (
    AREA_OTHER_HEX,
    area_palette_hex,
    calendar_display_compact_slot_label,
    calendar_display_slot_area,
    calendar_display_slot_sort_key,
    is_review_slot,
)


_BASE_COLS = [
    "day", "franja", "slot_id", "presentiality", "work_mode",
]

# Estils visuals (fons + colors de text consistents amb el PDF).
_FILL_HEADER = PatternFill("solid", fgColor="334155")
_FILL_AREA_OTHER = PatternFill("solid", fgColor=AREA_OTHER_HEX.lstrip("#"))
_FILL_PEONADA = PatternFill("solid", fgColor="FFF0C2")
_FILL_FESTIU = PatternFill("solid", fgColor="E5E7EB")
_FILL_LABEL = PatternFill("solid", fgColor="F1F5F9")
_BORDER_THIN = Border(*[Side(style="thin", color="CBD5E1")] * 4)
_FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_LABEL = Font(name="Calibri", size=9, bold=True, color="334155")
_FONT_PRES = Font(name="Calibri", size=9, color="15803D")
_FONT_NP = Font(name="Calibri", size=9, color="1D4ED8")
_FONT_MIXED = Font(name="Calibri", size=9, color="334155")
_FONT_GUARD = Font(name="Calibri", size=9, bold=True, color="B91C1C")
_FONT_DEFAULT = Font(name="Calibri", size=9, color="334155")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
# Nom de màquina en VERTICAL (com al PDF): així un nom llarg de
# qualsevol usuari hi cap sempre, per estreta que sigui la subcolumna.
_ALIGN_VERTICAL = Alignment(
    horizontal="center", vertical="center", text_rotation=90,
)


def _area_fill(area: str) -> PatternFill:
    # Paleta única a domain (schedule_format.area_palette_hex).
    return PatternFill("solid", fgColor=area_palette_hex(area).lstrip("#"))


def _read_csv_safe(path: Path) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError):
        return pd.DataFrame()


def _is_reforc(kind: str) -> bool:
    return str(kind or "").strip().lower().startswith(("refuerzo", "reforc", "reforç"))


def _guard_index(year: int | None, months: list[int] | None) -> dict:
    """Indexa guàrdies per dia (YYYY-MM-DD): {day: [(prof, kind)]}."""
    if year is None:
        return {}
    df = _read_csv_safe(Path("data/guards/assignments.csv"))
    if df.empty:
        return {}
    df = df.copy()
    df["_day"] = pd.to_datetime(df["day"], errors="coerce")
    df = df.dropna(subset=["_day"])
    out: dict = {}
    for _, row in df.iterrows():
        d = row["_day"].strftime("%Y-%m-%d")
        out.setdefault(d, []).append(
            (str(row["professional_id"]), str(row["guard_kind"]))
        )
    return out


def _abs_index(year: int | None, months: list[int] | None) -> dict:
    """Indexa absències per dia (YYYY-MM-DD): {day: [prof,...]}."""
    df = _read_csv_safe(Path("data/absences/assignments.csv"))
    if df.empty:
        return {}
    df = df.copy()
    df["_start"] = pd.to_datetime(df.get("start_day"), errors="coerce")
    df["_end"] = pd.to_datetime(df.get("end_day"), errors="coerce")
    df = df.dropna(subset=["_start", "_end"])
    out: dict = {}
    for _, row in df.iterrows():
        days = pd.date_range(row["_start"], row["_end"], freq="D")
        for d in days:
            out.setdefault(d.strftime("%Y-%m-%d"), []).append(
                str(row["professional_id"])
            )
    return out


def _post_guard_days(year: int | None) -> dict:
    """Dia → llista de facultatius en postguàrdia. Es deriva: post[d+1] = guards[d]."""
    if year is None:
        return {}
    df = _read_csv_safe(Path(f"data/derived/guard_constraints_{year}.csv"))
    if df.empty or "constraint_type" not in df.columns:
        return {}
    df = df[df["constraint_type"] == "post_guard_free"].copy()
    df["_day"] = pd.to_datetime(df["day"], errors="coerce")
    out: dict = {}
    for _, row in df.dropna(subset=["_day"]).iterrows():
        d = row["_day"].strftime("%Y-%m-%d")
        out.setdefault(d, []).append(str(row["professional_id"]))
    return out


_WEEKDAY_LABELS_CA = ["Dl", "Dm", "Dc", "Dj", "Dv"]


def _scope_title(year: int | None, months: list[int] | None) -> str:
    """Etiqueta del període per al títol (p.ex. «juny 2026» o
    «juny–juliol 2026»). Buida si no hi ha any."""
    from src.domain.month_scope import catalan_month_name
    if not year:
        return ""
    ms = [int(m) for m in (months or []) if 1 <= int(m) <= 12]
    if not ms:
        return str(year)
    if len(ms) == 1:
        return f"{catalan_month_name(ms[0])} {year}"
    return f"{catalan_month_name(ms[0])}–{catalan_month_name(ms[-1])} {year}"


def _cell_text_and_style(rows: pd.DataFrame) -> tuple[str, Font, PatternFill | None]:
    """Donat un DataFrame amb les files assignades a (day, franja, slot),
    retorna (text, font, fill_bg)."""
    if rows.empty:
        return "", _FONT_DEFAULT, None
    # Ordena: no-flipats primer, després flipats (perquè T- surt al segon).
    rows = rows.copy()
    rows["_flip"] = pd.to_numeric(rows.get("is_flipped", 0), errors="coerce").fillna(0).astype(int)
    rows = rows.sort_values(["_flip", "professional"]).reset_index(drop=True)
    parts = []
    has_pres = False
    has_np = False
    has_peonada = False
    for _, r in rows.iterrows():
        name = str(r.get("professional") or "").strip()
        if not name:
            continue
        if int(r["_flip"]) == 1:
            name = f"T-{name}"
        parts.append(name)
        pres = str(r.get("presentiality", "")).upper()
        if pres == "PRESENCIAL":
            has_pres = True
        elif pres == "NO_PRESENCIAL":
            has_np = True
        if str(r.get("work_mode", "")).upper() == "PEONADA":
            has_peonada = True
    text = "/".join(parts)
    if has_pres and not has_np:
        font = _FONT_PRES
    elif has_np and not has_pres:
        font = _FONT_NP
    else:
        font = _FONT_MIXED
    fill = _FILL_PEONADA if has_peonada else None
    return text, font, fill


def _area_sort_key(area: str) -> tuple[int, str]:
    """Ordre de les àrees: alfabètic, amb 'ALTRES' (sense àrea) al final."""
    a = str(area).strip().upper()
    return (1, "") if (not a or a == "ALTRES") else (0, a)


def _machines_for_day(week_df: pd.DataFrame, dstr: str) -> list[str]:
    """Llista ordenada (per àrea i slot) de màquines amb assignació un dia."""
    sub = week_df[week_df["day"] == dstr]
    if sub.empty:
        return []
    machines = sorted(set(sub["_sid"].tolist()))
    return sorted(
        machines,
        key=lambda m: (
            _area_sort_key(calendar_display_slot_area(m)),
            calendar_display_slot_sort_key(m),
            m,
        ),
    )


def _prepared_week_df(schedule_df: pd.DataFrame, days_str: list[str]) -> pd.DataFrame:
    """Files de màquina del bloc (sense guàrdies ni files sense
    facultatiu), amb les columnes auxiliars _sid/_pid/_franja."""
    week_df = schedule_df[schedule_df["day"].isin(days_str)].copy()
    if "is_flipped" not in week_df.columns:
        week_df["is_flipped"] = 0
    week_df["_sid"] = week_df["slot_id"].astype(str).str.strip().str.upper()
    week_df["_pid"] = week_df["professional"].astype(str).str.strip().str.upper()
    week_df["_franja"] = week_df["franja"].astype(str).str.strip().str.upper()
    return week_df[
        ~week_df["_sid"].isin(GUARDS_RESERVED_SLOT_IDS)
        & ~week_df["_pid"].isin({"", "NONE", "NAN"})
        & (week_df["_franja"].isin({"MATI", "TARDA", "NIT"}))
    ].copy()


def _machine_col_ranges(c0: int, n_sub: int, k: int) -> list[tuple[int, int]]:
    """Reparteix les `n_sub` subcolumnes del bloc d'un dia entre les
    seves `k` màquines (com el PDF: el dia sempre ocupa el mateix
    ample i les màquines se'l parteixen a parts iguals)."""
    out = []
    for i in range(k):
        cs = c0 + round(i * n_sub / k)
        ce = c0 + round((i + 1) * n_sub / k) - 1
        out.append((cs, max(cs, ce)))
    return out


def _merged_cell(ws, row, c0, c1, value, font, fill=None,
                 align=_ALIGN_CENTER):
    """Cel·la (potser fusionada) amb estil aplicat a TOT el rang, perquè
    les vores i el fons cobreixin el bloc sencer."""
    for cc in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=cc)
        cell.border = _BORDER_THIN
        if fill is not None:
            cell.fill = fill
    cell = ws.cell(row=row, column=c0, value=value)
    cell.font = font
    cell.alignment = align
    cell.border = _BORDER_THIN
    if fill is not None:
        cell.fill = fill
    if c1 > c0:
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
    return cell


def _write_week_block(
    ws,
    week_days: list[pd.Timestamp],
    schedule_df: pd.DataFrame,
    guards_idx: dict,
    abs_idx: dict,
    post_idx: dict,
    festiu_days: set[str],
    start_row: int,
    n_sub: int = 1,
) -> int:
    """Escriu un bloc setmanal a `ws` amb el layout del PDF setmanal:
    TOTS els dies ocupen el mateix ample (`n_sub` subcolumnes) i les
    màquines de cada dia es reparteixen el bloc; les files són
    [Header / Abs / Àrea / Màquina / Matí / Tarda / Nit]. Retorna la
    fila següent disponible després del bloc."""
    days_str = [d.strftime("%Y-%m-%d") for d in week_days]
    week_df = _prepared_week_df(schedule_df, days_str)

    # Slots de revisió (dia sencer): files separades sota les franges.
    review_rows = schedule_df[
        schedule_df["day"].isin(days_str)
        & schedule_df["slot_id"].astype(str).map(is_review_slot)
        & ~schedule_df["professional"].astype(str).str.upper().isin({"", "NONE", "NAN"})
    ].copy()

    # Màquines per dia (només dies del bloc; festius queden sense
    # màquines però mantenen el bloc sencer).
    machines_per_day = {d: _machines_for_day(week_df, d.strftime("%Y-%m-%d")) for d in week_days}

    # Blocs de dia d'AMPLE FIX: cada dia ocupa exactament n_sub
    # subcolumnes (com al PDF), independentment de quantes màquines
    # tingui — així tots els dies queden alineats entre setmanes.
    n_sub = max(1, int(n_sub), *(len(m) for m in machines_per_day.values()))
    day_ranges: dict = {}
    machine_ranges: dict = {}
    col = 2  # col 1 = etiqueta de fila
    for d in week_days:
        day_ranges[d] = (col, col + n_sub - 1)
        machine_ranges[d] = _machine_col_ranges(
            col, n_sub, max(1, len(machines_per_day[d])),
        )
        col += n_sub
    total_cols = col - 1
    n_cols = total_cols  # per a estendre fills

    def _set_label(r: int, label: str) -> None:
        c = ws.cell(row=r, column=1, value=label)
        c.fill = _FILL_LABEL
        c.font = _FONT_LABEL
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_THIN

    # ── Row 1: capçaleres dels dies (amb guàrdia/reforç en vermell).
    r = start_row
    # Etiqueta buida a la col 1.
    head_label = ws.cell(row=r, column=1, value="")
    head_label.fill = _FILL_HEADER
    head_label.border = _BORDER_THIN
    for d in week_days:
        c0, c1 = day_ranges[d]
        dstr = d.strftime("%Y-%m-%d")
        label = f"{_WEEKDAY_LABELS_CA[d.weekday()]} {d.day}"
        g = guards_idx.get(dstr) or []
        if g:
            gtxt = " · " + " ".join(
                f"{p} ({'R' if _is_reforc(k) else 'G'})" for p, k in g
            )
            label = label + gtxt
        _merged_cell(
            ws, r, c0, c1, label,
            _FONT_GUARD if g else _FONT_HEADER, fill=_FILL_HEADER,
        )
    r += 1

    # ── Row 2: absències (per dia, amb postguàrdia (PG)).
    _set_label(r, "Abs.")
    for d in week_days:
        c0, c1 = day_ranges[d]
        dstr = d.strftime("%Y-%m-%d")
        items = list(abs_idx.get(dstr, []))
        items += [f"{p} (PG)" for p in post_idx.get(dstr, [])]
        _merged_cell(
            ws, r, c0, c1, " ".join(items), _FONT_DEFAULT,
            fill=_FILL_FESTIU if dstr in festiu_days else None,
        )
    r += 1

    # ── Row 3: ÀREA. Per a cada dia, fusiona màquines consecutives
    #          de la mateixa àrea amb el fons corresponent.
    _set_label(r, "Àrea")
    for d in week_days:
        c0, c1 = day_ranges[d]
        machines = machines_per_day[d]
        if not machines:
            # Bloc sencer buit (festiu o dia sense màquines).
            _merged_cell(
                ws, r, c0, c1, "", _FONT_DEFAULT,
                fill=_FILL_FESTIU
                if d.strftime("%Y-%m-%d") in festiu_days else None,
            )
            continue
        # Agrupa per àrea (consecutives), sobre els rangs de màquina.
        ranges = machine_ranges[d]
        i = 0
        while i < len(machines):
            cur_area = calendar_display_slot_area(machines[i])
            j = i
            while j < len(machines) and calendar_display_slot_area(machines[j]) == cur_area:
                j += 1
            _merged_cell(
                ws, r, ranges[i][0], ranges[j - 1][1], cur_area,
                _FONT_LABEL, fill=_area_fill(cur_area),
            )
            i = j
    r += 1

    # ── Row 4: MÀQUINA (una etiqueta per màquina, sobre el seu rang).
    _set_label(r, "Màquina")
    for d in week_days:
        c0, c1 = day_ranges[d]
        machines = machines_per_day[d]
        dstr = d.strftime("%Y-%m-%d")
        if not machines:
            _merged_cell(
                ws, r, c0, c1, "", _FONT_DEFAULT,
                fill=_FILL_FESTIU if dstr in festiu_days else None,
            )
            continue
        for (ms, me), m in zip(machine_ranges[d], machines):
            # Etiqueta COMPACTA (com al PDF): el sufix del lloc es treu
            # — l'àrea ja surt a la fila de sobre (ECO1_LLOC → ECO1) —
            # i el text va en VERTICAL perquè hi càpiga qualsevol nom.
            _merged_cell(
                ws, r, ms, me, calendar_display_compact_slot_label(m),
                _FONT_LABEL,
                fill=_area_fill(calendar_display_slot_area(m)),
                align=_ALIGN_VERTICAL,
            )
    # Registrem la fila de Màquina: en fixar alçades li donarem més
    # espai (els noms van en vertical, com la franja alta del PDF).
    ws._machine_rows = getattr(ws, "_machine_rows", []) + [r]
    r += 1

    # ── Rows 5,6,7: Matí, Tarda, Nit. Una cel·la per (dia, màquina).
    for franja in ("MATI", "TARDA", "NIT"):
        _set_label(r, franja.title())
        for d in week_days:
            c0, c1 = day_ranges[d]
            dstr = d.strftime("%Y-%m-%d")
            machines = machines_per_day[d]
            if not machines:
                _merged_cell(
                    ws, r, c0, c1, "", _FONT_DEFAULT,
                    fill=_FILL_FESTIU if dstr in festiu_days else None,
                )
                continue
            for (ms, me), m in zip(machine_ranges[d], machines):
                rows = week_df[
                    (week_df["day"] == dstr)
                    & (week_df["_sid"] == m)
                    & (week_df["_franja"] == franja)
                ]
                text, font, fill = _cell_text_and_style(rows)
                if dstr in festiu_days and not text:
                    fill = _FILL_FESTIU
                _merged_cell(ws, r, ms, me, text, font, fill=fill)
        r += 1

    # ── Revisions: una fila per slot de revisió. Span per dia
    #              (s'aplica al dia sencer; mostrem el facultatiu
    #              fusionant les màquines del dia).
    review_slots = sorted(
        {str(rr.slot_id).strip().upper() for rr in review_rows.itertuples(index=False)},
        key=calendar_display_slot_sort_key,
    )
    for sid in review_slots:
        _set_label(r, f"Rev. {calendar_display_compact_slot_label(sid)}")
        for d in week_days:
            c0, c1 = day_ranges[d]
            dstr = d.strftime("%Y-%m-%d")
            rev_rows = review_rows[
                (review_rows["day"] == dstr)
                & (review_rows["slot_id"].astype(str).str.strip().str.upper() == sid)
            ]
            text, font, fill = _cell_text_and_style(rev_rows)
            if dstr in festiu_days and not text:
                fill = _FILL_FESTIU
            _merged_cell(ws, r, c0, c1, text, font, fill=fill)
        r += 1

    # Track max cols per a freeze i autosize aproximat
    ws._pdf_like_max_col = max(getattr(ws, "_pdf_like_max_col", 0), n_cols)
    return r + 1  # fila buida com a separador entre setmanes


def _non_working_days(year: int) -> set[str]:
    path = Path(f"data/base_calendar_{year}.csv")
    if not path.exists() or path.stat().st_size == 0:
        return set()
    df = pd.read_csv(path)
    if not {"day", "is_working_day"}.issubset(df.columns):
        return set()
    df["day"] = pd.to_datetime(df["day"], errors="coerce")
    df["is_working_day"] = pd.to_numeric(df["is_working_day"], errors="coerce").fillna(1).astype(int)
    return set(
        df.loc[df["is_working_day"] == 0, "day"]
        .dt.strftime("%Y-%m-%d")
    )


def _write_month_grid(
    book,
    sheet_name: str,
    df: pd.DataFrame,
    year: int | None,
    months: list[int] | None,
    guards_idx: dict,
    abs_idx: dict,
    post_idx: dict,
    festiu_days: set[str],
    index: int | None = None,
) -> None:
    """Escriu UNA fulla-quadrícula (un mes = una pàgina A4 apaïsada).
    `df` ja ve filtrat al període de la fulla."""
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    ws = (
        book.create_sheet(sheet_name, index)
        if index is not None else book.create_sheet(sheet_name)
    )

    # Conjunt de dilluns únics del període (les setmanes frontereres hi
    # entren pels seus dies del mes; els dies d'altres mesos queden buits).
    dfm = df.copy()
    dfm["_monday"] = dfm["_day_dt"] - pd.to_timedelta(
        dfm["_day_dt"].dt.weekday, unit="D"
    )
    mondays = sorted(set(dfm["_monday"].dt.normalize()))

    # Pre-passada: nombre de subcolumnes per dia (el MÀXIM de màquines
    # d'un dia en tot el full). Tots els dies ocupen aquest mateix ample
    # (com al PDF) i les màquines de cada dia se'l reparteixen — així
    # els dies queden alineats entre setmanes, festius inclosos.
    n_sub = 1
    for mon in mondays:
        week_days = [mon + pd.Timedelta(days=k) for k in range(5)]
        days_str = [d.strftime("%Y-%m-%d") for d in week_days]
        wdf = _prepared_week_df(dfm, days_str)
        for ds in days_str:
            n_sub = max(n_sub, len(_machines_for_day(wdf, ds)))

    # Reservem la fila 1 per al títol; els blocs setmanals comencen a la 2.
    current_row = 2
    for mon in mondays:
        week_days = [mon + pd.Timedelta(days=k) for k in range(5)]
        # Salta setmanes sense cap assignació.
        days_str = {d.strftime("%Y-%m-%d") for d in week_days}
        if not (dfm["day"].isin(days_str)).any():
            continue
        current_row = _write_week_block(
            ws,
            week_days,
            dfm,
            guards_idx,
            abs_idx,
            post_idx,
            festiu_days,
            current_row,
            n_sub=n_sub,
        )

    max_col = max(2, getattr(ws, "_pdf_like_max_col", 2))

    # Amplada de columnes NOMÉS per a les realment usades: donar amplada
    # a columnes buides (l'antic range(1, 50)) les incorporava a l'àrea
    # d'impressió d'Excel i l'escala «ajusta a pàgina» quedava
    # esguerrada. El bloc de cada dia té un ample total ~constant (les
    # subcolumnes es fan més estretes com més màquines hi ha).
    ws.column_dimensions["A"].width = 10
    sub_width = min(15.0, max(3.5, round(45.0 / max(1, n_sub), 1)))
    for i in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = sub_width

    # ── Títol (fila 1), fusionat sobre tota l'amplada del calendari.
    title = "Calendari entre setmana"
    period = _scope_title(year, months)
    if period:
        title = f"{title} · {period}"
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    if max_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.row_dimensions[1].height = 30

    # Alçada de fila CALCULADA perquè la proporció del contingut
    # coincideixi amb l'A4 APAÏSAT: l'ajusta-a-pàgina usa una escala
    # uniforme, i si el full és proporcionalment més alt que ample
    # (mesos amb moltes setmanes/files) queden marges laterals enormes;
    # si és més ample, queda arraconat a dalt. Amplada en punts ≈
    # cols × ample_caràcter (~5.25 pt/char); objectiu = àrea útil de
    # l'A4 horitzontal (~11,1 × 7,5 polzades).
    n_content_rows = max(1, current_row - 2)
    machine_rows = set(getattr(ws, "_machine_rows", []))
    # Les files de Màquina (noms en vertical) són ~2,4× més altes, com
    # la franja de màquines del PDF; l'equació de proporció ho incorpora.
    _MACH_FACTOR = 2.4
    eff_rows = n_content_rows + (len(machine_rows) * (_MACH_FACTOR - 1))
    width_pt = (10 + (max_col - 1) * sub_width) * 5.25
    target_height_pt = width_pt * 7.5 / 11.1
    row_h = min(34.0, max(10.0, (target_height_pt - 30) / eff_rows))
    for rr in range(2, current_row):
        ws.row_dimensions[rr].height = (
            row_h * _MACH_FACTOR if rr in machine_rows else row_h
        )

    # Congela el títol i la columna d'etiquetes.
    ws.freeze_panes = "B2"

    # Impressió: HORITZONTAL (landscape) en A4, ajustada a UNA pàgina
    # (amplada i alçada), centrada, i amb l'ÀREA D'IMPRESSIÓ explícita
    # (només les cel·les usades — res de columnes fantasma).
    last_row = max(2, current_row - 1)
    ws.print_area = f"A1:{get_column_letter(max_col)}{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2,
    )


def _write_calendari_grid(
    writer,
    schedule_df: pd.DataFrame,
    year: int | None,
    months: list[int] | None,
) -> None:
    """Escriu la quadrícula setmanal (mimica PDF). Amb UN mes al scope,
    una única fulla «Calendari»; amb diversos, UNA FULLA PER MES — cada
    mes imprimeix en la seva pròpia pàgina A4 apaïsada (abans tot el
    scope s'encabia en una sola pàgina i l'escala variava segons els
    festius i les setmanes sense programació)."""
    # Carrega dades operatives (per a Abs i Guàrdia a la capçalera)
    guards_idx = _guard_index(year, months)
    abs_idx = _abs_index(year, months)
    post_idx = _post_guard_days(year) if year else {}
    festiu_days = _non_working_days(year) if year else set()

    df = schedule_df.copy()
    df["day"] = df["day"].astype(str)
    df["_day_dt"] = pd.to_datetime(df["day"], errors="coerce")
    if year is not None and months:
        from src.domain.month_scope import in_logical_months
        df = df[in_logical_months(df["_day_dt"], year, months)]
    df = df.dropna(subset=["_day_dt"]).copy()
    if df.empty:
        ws = writer.book.create_sheet("Calendari", 0)
        ws.cell(row=1, column=1, value="(sense assignacions per al scope)")
        return

    month_list = sorted({int(m) for m in months}) if (year and months) else []
    if len(month_list) > 1:
        from src.domain.month_scope import catalan_month_name, in_logical_months
        idx = 0
        for m in month_list:
            sub = df[in_logical_months(df["_day_dt"], year, [m])].copy()
            if sub.empty:
                continue
            _write_month_grid(
                writer.book, f"Calendari {catalan_month_name(m)}", sub,
                year, [m], guards_idx, abs_idx, post_idx, festiu_days,
                index=idx,
            )
            idx += 1
    else:
        _write_month_grid(
            writer.book, "Calendari", df, year, months,
            guards_idx, abs_idx, post_idx, festiu_days, index=0,
        )


def _register_catalog_overrides() -> None:
    """Registra els overrides del catàleg (àrea/família/revisió) perquè
    la classificació de slots a l'Excel coincideixi amb la del PDF."""
    try:
        from src.services.slot_catalog import (
            load_slot_catalog, slot_area_map,
            slot_metric_family_map, review_slot_ids,
        )
        from src.domain.schedule_format import (
            set_slot_area_overrides, set_slot_metric_overrides,
            set_slot_review_overrides,
        )
        cat_path = Path("data/slot_catalog.csv")
        if cat_path.exists():
            cat = load_slot_catalog(cat_path)
            set_slot_area_overrides(slot_area_map(cat))
            set_slot_metric_overrides(slot_metric_family_map(cat))
            set_slot_review_overrides(review_slot_ids(cat))
    except (OSError, ImportError, AttributeError):
        pass


def export_schedule_to_excel(
    schedule_csv: Path,
    output_xlsx: Path,
    selected_months: list[int] | None = None,
    year: int | None = None,
) -> int:
    """Genera l'Excel amb la quadrícula PDF-like + fulles auxiliars.
    Retorna el nombre de files de l'schedule processades."""
    _register_catalog_overrides()
    schedule_csv = Path(schedule_csv)
    output_xlsx = Path(output_xlsx)
    if not schedule_csv.exists() or schedule_csv.stat().st_size == 0:
        return 0
    df = pd.read_csv(schedule_csv)
    if df.empty:
        return 0

    if year is not None or selected_months:
        from src.domain.month_scope import in_logical_months
        df = df.copy()
        df["_day_dt"] = pd.to_datetime(df["day"], errors="coerce")
        if year is not None and selected_months:
            df = df[in_logical_months(df["_day_dt"], year, selected_months)]
        elif year is not None:
            df = df[df["_day_dt"].dt.year == year]
        df = df.drop(columns=["_day_dt"])
        if df.empty:
            return 0

    cols_present = [c for c in _BASE_COLS if c in df.columns]
    df = df.sort_values(cols_present).reset_index(drop=True)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        # Única fulla «Calendari» amb la quadrícula PDF-like.
        _write_calendari_grid(writer, df, year, selected_months)
    return int(len(df))
